#!/usr/bin/env python3
"""
Apply Stage 1 source and human-decision tags to Zotero items based on current
collection membership, and generate an overlap report.

Tag schema applied:
  source:<key>           — for every screening collection the item appears in
                           (one tag per database; multi-tag if cross-source dup)
  s1:human:<decision>    — for non-SSRN items, derived from current
                           01-Keep / 02-Maybe / 03-Discard membership
                           (no tag for items in 00-Queue)
                           (SSRN items get NO decision tag here — those come
                           later when Claude/ChatGPT/Gemini results import)

Overlap report (XLSX):
  Sheet 1 - Summary       : counts by overlap pattern
  Sheet 2 - Cross-DB      : per-item detail of items in 2+ source databases
  Sheet 3 - Superseded    : per-item detail of items linked to Superseded
  Sheet 4 - All overlaps  : per-item detail of any item in 2+ tracked collections

Usage:
    python3 apply_stage1_tags.py             # dry-run: report only, no writes
    python3 apply_stage1_tags.py --apply     # actually write tags
    python3 apply_stage1_tags.py --report-only   # skip tag analysis, just overlap report

Output:
    overlap_report.xlsx       — always written
    tag_plan.csv              — what would change (dry-run) or did change (apply)
    apply.log                 — execution log
"""
import json
import urllib.request
import urllib.error
import csv
import os
import sys
import time
import random
import argparse
from datetime import datetime, timezone
from collections import defaultdict
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

# ============================================================
# CONFIG
# ============================================================
ZOTERO_API_KEY_RO = os.environ.get("ZOTERO_API_KEY_RO", "")  # read-only
ZOTERO_API_KEY_RW = os.environ.get("ZOTERO_API_KEY_RW", "")  # read-write (used only with --apply)

LIB = "6505702"
BASE = f"https://api.zotero.org/groups/{LIB}"
RATE_LIMIT_SEC = 0.3

# 02-Screening parent collection keys per source
SCREENING_PARENTS = {
    "IEEE Xplore":                   ("7XHWH8NM", "ieee"),
    "SCOPUS":                        ("2RWBC7QH", "scopus"),
    "Web of Science":                ("E7AS4HD4", "wos"),
    "arXiv":                         ("YK2CHQLN", "arxiv"),
    "ACM":                           ("G4IIYGV6", "acm"),
    "Practitioner Network":          ("7FL4M8HN", "practitioner"),
    "Coursework":                    ("IF299TAY", "coursework"),
    "Hancheng Cao":                  ("GP3U9EX8", "cao"),
    "Linda Naimi Book - References": ("ZXIXRWKG", "naimi-references"),
    "Linda Naimi Book - Chapters":   ("DIDV7BKH", "naimi-chapters"),
    "SSRN":                          ("F9A9883N", "ssrn"),
}

# Decision child collection name patterns -> tag suffix
DECISION_BUCKETS = {
    "01-Keep":    "keep",
    "02-Maybe":   "maybe",
    "03-Discard": "discard",
    # 00-Queue intentionally absent — no decision tag
}

OUTPUT_REPORT = "overlap_report.xlsx"
TAG_PLAN_CSV = "tag_plan.csv"
LOG_FILE = "apply.log"

# ============================================================
# LOGGING
# ============================================================
def log(msg):
    line = f"[{datetime.now(timezone.utc).strftime('%Y-%m-%d %H:%M:%S')}] {msg}"
    print(line, flush=True)
    with open(LOG_FILE, "a") as f:
        f.write(line + "\n")

# ============================================================
# ZOTERO API
# ============================================================
def zot_get(path, retries=8, api_key=None):
    last_err = None
    for attempt in range(retries):
        try:
            req = urllib.request.Request(f"{BASE}{path}")
            req.add_header("Zotero-API-Key", api_key or ZOTERO_API_KEY_RO)
            req.add_header("Zotero-API-Version", "3")
            with urllib.request.urlopen(req, timeout=60) as r:
                return json.loads(r.read()), {k.lower(): v for k, v in r.headers.items()}
        except urllib.error.HTTPError as e:
            if e.code in (429, 500, 502, 503, 504) and attempt < retries - 1:
                wait = min(60, 2 ** attempt) + random.uniform(0, 1)
                log(f"  HTTP {e.code} on {path}, retry in {wait:.1f}s")
                time.sleep(wait)
                continue
            last_err = f"HTTP {e.code}"
        except (urllib.error.URLError, TimeoutError, ConnectionError) as e:
            if attempt < retries - 1:
                wait = min(60, 2 ** attempt) + random.uniform(0, 1)
                log(f"  Network error on {path}: {e}; retry in {wait:.1f}s")
                time.sleep(wait)
                continue
            last_err = str(e)
        except Exception as e:
            last_err = str(e)
            if attempt < retries - 1:
                time.sleep(2 ** attempt)
                continue
    return None, {"error": last_err}

def zot_patch(path, data, version, api_key):
    """PATCH an item with new fields (e.g. tags). Requires write key and current version."""
    body = json.dumps(data).encode("utf-8")
    for attempt in range(8):
        try:
            req = urllib.request.Request(f"{BASE}{path}", data=body, method="PATCH")
            req.add_header("Zotero-API-Key", api_key)
            req.add_header("Zotero-API-Version", "3")
            req.add_header("Content-Type", "application/json")
            req.add_header("If-Unmodified-Since-Version", str(version))
            with urllib.request.urlopen(req, timeout=60) as r:
                # Empty body on success (204)
                return True, None
        except urllib.error.HTTPError as e:
            if e.code in (429, 500, 502, 503, 504) and attempt < 7:
                wait = min(60, 2 ** attempt) + random.uniform(0, 1)
                log(f"  PATCH HTTP {e.code}, retry in {wait:.1f}s")
                time.sleep(wait)
                continue
            err_body = ""
            try:
                err_body = e.read().decode("utf-8", errors="replace")[:300]
            except Exception:
                pass
            return False, f"HTTP {e.code}: {err_body}"
        except Exception as e:
            if attempt < 7:
                time.sleep(2 ** attempt)
                continue
            return False, str(e)
    return False, "max retries"

# ============================================================
# COLLECTION DISCOVERY
# ============================================================
def all_collections():
    """Return list of all collections in the library: [{key, name, parent}, ...]."""
    log("Scanning all collections in library...")
    out = []
    start = 0
    while True:
        data, hdrs = zot_get(f"/collections?limit=100&start={start}")
        if data is None:
            log(f"  ERROR: {hdrs.get('error','?')}")
            sys.exit(1)
        if not data:
            break
        for c in data:
            d = c.get("data", {})
            out.append({
                "key": c["key"],
                "name": d.get("name", ""),
                "parent": d.get("parentCollection", "") or "",
            })
        if len(data) < 100:
            break
        start += 100
        time.sleep(RATE_LIMIT_SEC)
    log(f"  Found {len(out)} collections total")
    return out

def descendants_of(parent_key, by_parent):
    """Return set of all descendant collection keys (any depth) under parent_key."""
    out = set()
    stack = [parent_key]
    while stack:
        k = stack.pop()
        for child in by_parent.get(k, []):
            if child not in out:
                out.add(child)
                stack.append(child)
    return out

def get_collection_items(collection_key):
    items = []
    start = 0
    while True:
        data, hdrs = zot_get(f"/collections/{collection_key}/items/top?limit=100&start={start}")
        if data is None:
            log(f"  ERROR fetching items of {collection_key}: {hdrs.get('error','?')}")
            return None
        if not data:
            break
        items.extend(data)
        if len(data) < 100:
            break
        start += 100
        time.sleep(RATE_LIMIT_SEC)
    return items

# ============================================================
# CORE: build classification map
# ============================================================
def build_classification(collections):
    """
    Walk every screening collection across all sources. For each item, record:
      - source memberships (set of source short-names)
      - decision per source (e.g., {"ieee": "keep", "scopus": "maybe"})
      - whether linked to any Superseded collection
    Also fetches item metadata (title, version, current tags) for use in tag application.

    Returns dict[item_key] = {
        "title": str, "version": int, "current_tags": [str,...],
        "sources": set[str], "decisions": dict[source_short -> decision_word],
        "superseded": bool, "superseded_collections": [name,...]
    }
    """
    by_key = {c["key"]: c for c in collections}
    by_parent = defaultdict(list)
    for c in collections:
        if c["parent"]:
            by_parent[c["parent"]].append(c["key"])

    # Identify superseded collection keys (any name containing 'supersed', case-insensitive)
    superseded_keys = {c["key"]: c["name"] for c in collections if "supersed" in c["name"].lower()}
    if superseded_keys:
        log(f"Superseded collections found: {len(superseded_keys)}")
        for k, n in superseded_keys.items():
            log(f"  {k}  {n}")
    else:
        log("No Superseded collections found")

    # For each source: find its decision-bucket child collection keys under the 02-Screening parent
    # screening_map: source_short -> dict[bucket_name -> collection_key]
    screening_map = {}
    queue_keys_by_source = {}  # source_short -> queue collection key (no decision tag, but still tracked for source membership)
    log("\nResolving 02-Screening child buckets per source...")
    for src_name, (parent_key, src_short) in SCREENING_PARENTS.items():
        # Direct children of this 02-Screening parent
        children = [by_key[k] for k in by_parent.get(parent_key, []) if k in by_key]
        bucket_keys = {}
        for c in children:
            for bucket_name in DECISION_BUCKETS:
                if c["name"].strip() == bucket_name:
                    bucket_keys[bucket_name] = c["key"]
            if c["name"].strip() == "00-Queue":
                queue_keys_by_source[src_short] = c["key"]
        screening_map[src_short] = bucket_keys
        log(f"  {src_name} ({src_short}): " + ", ".join(f"{b}={k}" for b, k in bucket_keys.items()))
        if src_short in queue_keys_by_source:
            log(f"    + 00-Queue={queue_keys_by_source[src_short]}")

    # Build the classification map
    classification = {}  # item_key -> dict as above

    def ensure(key, item):
        if key not in classification:
            d = item.get("data", {})
            classification[key] = {
                "title": d.get("title", "") or "",
                "version": item.get("version", 0),
                "current_tags": [t.get("tag", "") for t in d.get("tags", [])],
                "sources": set(),
                "decisions": {},          # source_short -> "keep"/"maybe"/"discard"
                "queue_only_in": set(),   # sources where item is in Queue (no decision)
                "superseded": False,
                "superseded_collections": [],
            }
        return classification[key]

    # Walk each source's bucket collections
    for src_name, (parent_key, src_short) in SCREENING_PARENTS.items():
        # Decision buckets
        for bucket_name, bucket_key in screening_map[src_short].items():
            decision = DECISION_BUCKETS[bucket_name]
            log(f"Fetching {src_name} / {bucket_name}...")
            items = get_collection_items(bucket_key)
            if items is None:
                continue
            log(f"  {len(items)} items")
            for it in items:
                k = it["key"]
                rec = ensure(k, it)
                rec["sources"].add(src_short)
                rec["decisions"][src_short] = decision
            time.sleep(RATE_LIMIT_SEC)
        # Queue (source membership only, no decision)
        qk = queue_keys_by_source.get(src_short)
        if qk:
            log(f"Fetching {src_name} / 00-Queue...")
            items = get_collection_items(qk)
            if items is None:
                continue
            log(f"  {len(items)} items")
            for it in items:
                k = it["key"]
                rec = ensure(k, it)
                rec["sources"].add(src_short)
                if src_short not in rec["decisions"]:
                    rec["queue_only_in"].add(src_short)
            time.sleep(RATE_LIMIT_SEC)

    # Mark superseded membership using each item's collections array
    log("\nChecking superseded membership...")
    superseded_count = 0
    for k, rec in classification.items():
        # We need the full collections list per item — fetch fresh? No, we have it from earlier.
        # But ensure() only stored what was in the 'data' for the response — collections array is there.
        pass  # We'll compute superseded in a second pass below using item-level fetches

    # Instead of re-fetching all items, walk each Superseded collection and mark
    for sk, sname in superseded_keys.items():
        log(f"Fetching superseded collection {sname}...")
        items = get_collection_items(sk)
        if items is None:
            continue
        log(f"  {len(items)} items")
        for it in items:
            k = it["key"]
            if k in classification:
                rec = classification[k]
                rec["superseded"] = True
                if sname not in rec["superseded_collections"]:
                    rec["superseded_collections"].append(sname)
            else:
                # Item is in a superseded collection but not in any screening collection.
                # Track minimally so the report can show it.
                d = it.get("data", {})
                classification[k] = {
                    "title": d.get("title", "") or "",
                    "version": it.get("version", 0),
                    "current_tags": [t.get("tag", "") for t in d.get("tags", [])],
                    "sources": set(),
                    "decisions": {},
                    "queue_only_in": set(),
                    "superseded": True,
                    "superseded_collections": [sname],
                }
        time.sleep(RATE_LIMIT_SEC)

    superseded_count = sum(1 for r in classification.values() if r["superseded"])
    log(f"  Items linked to superseded collections: {superseded_count}")

    return classification

# ============================================================
# TAG PLAN
# ============================================================
def build_tag_plan(classification):
    """
    For each item, compute the desired set of tags to ADD (not remove).
    Returns list of dicts: {key, title, current_tags, desired_new_tags, sources, decisions, superseded}
    """
    plan = []
    for k, rec in classification.items():
        desired = set()
        # Source tags for every source membership (decision or queue)
        for src in rec["sources"]:
            desired.add(f"source:{src}")
        # Decision tags only for non-SSRN items
        for src, decision in rec["decisions"].items():
            if src == "ssrn":
                continue  # Per requirement: SSRN gets source only, decisions come later
            desired.add(f"s1:human:{decision}")
        # Subtract tags already present
        current = set(rec["current_tags"])
        new_tags = sorted(desired - current)
        if new_tags or rec["sources"]:  # include in plan even if no changes (for visibility)
            plan.append({
                "key": k,
                "title": rec["title"],
                "current_tags": sorted(current),
                "desired_new_tags": new_tags,
                "all_desired_tags": sorted(desired),
                "sources": sorted(rec["sources"]),
                "decisions": dict(rec["decisions"]),
                "queue_only_in": sorted(rec["queue_only_in"]),
                "superseded": rec["superseded"],
                "superseded_collections": rec["superseded_collections"],
                "version": rec["version"],
            })
    return plan

def write_tag_plan_csv(plan, path):
    with open(path, "w", newline="", encoding="utf-8") as f:
        w = csv.writer(f)
        w.writerow([
            "item_key", "title", "sources", "decisions",
            "queue_only_in", "superseded", "superseded_collections",
            "current_tags", "tags_to_add", "all_desired_tags"
        ])
        for p in plan:
            w.writerow([
                p["key"], p["title"][:200],
                "|".join(p["sources"]),
                "|".join(f"{s}:{d}" for s, d in sorted(p["decisions"].items())),
                "|".join(p["queue_only_in"]),
                "yes" if p["superseded"] else "",
                "|".join(p["superseded_collections"]),
                "|".join(p["current_tags"]),
                "|".join(p["desired_new_tags"]),
                "|".join(p["all_desired_tags"]),
            ])

# ============================================================
# TAG APPLICATION
# ============================================================
def apply_tags(plan, api_key):
    """Apply tags via PATCH for each item with new tags to add."""
    to_apply = [p for p in plan if p["desired_new_tags"]]
    log(f"\nApplying tags to {len(to_apply)} items...")
    successes = 0
    failures = []
    for i, p in enumerate(to_apply, 1):
        # Build the new tags array: union of existing + new
        all_tags = sorted(set(p["current_tags"]) | set(p["desired_new_tags"]))
        tag_objects = [{"tag": t} for t in all_tags]
        ok, err = zot_patch(
            f"/items/{p['key']}",
            {"tags": tag_objects},
            p["version"],
            api_key,
        )
        if ok:
            successes += 1
            if i % 50 == 0 or i == len(to_apply):
                log(f"  [{i}/{len(to_apply)}] applied (success={successes}, fail={len(failures)})")
        else:
            failures.append((p["key"], err))
            log(f"  FAIL {p['key']}: {err}")
        time.sleep(RATE_LIMIT_SEC)
    log(f"\nApply complete. Success: {successes}, Failures: {len(failures)}")
    if failures:
        log("Failures:")
        for k, e in failures[:20]:
            log(f"  {k}: {e}")
        if len(failures) > 20:
            log(f"  ... and {len(failures)-20} more")
    return successes, failures

# ============================================================
# OVERLAP REPORT (XLSX)
# ============================================================
def generate_overlap_report(classification, path):
    log(f"\nGenerating overlap report: {path}")
    
    wb = Workbook()
    
    # Styles
    hdr_font = Font(bold=True, color="FFFFFF")
    hdr_fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
    section_fill = PatternFill(start_color="EBF3FA", end_color="EBF3FA", fill_type="solid")
    
    def style_header(ws, row, ncols):
        for col in range(1, ncols + 1):
            cell = ws.cell(row=row, column=col)
            cell.font = hdr_font
            cell.fill = hdr_fill
            cell.alignment = Alignment(horizontal="left", vertical="center")
    
    def autosize(ws, max_width=80):
        for col_cells in ws.columns:
            length = max(len(str(c.value or "")) for c in col_cells)
            col_letter = get_column_letter(col_cells[0].column)
            ws.column_dimensions[col_letter].width = min(max_width, max(10, length + 2))

    # ------------------------------------------------------------
    # Sheet 1: Summary
    # ------------------------------------------------------------
    ws = wb.active
    ws.title = "Summary"
    
    ws.append(["SLR Overlap Report"])
    ws.cell(row=1, column=1).font = Font(bold=True, size=14)
    ws.append([f"Generated: {datetime.now(timezone.utc).strftime('%Y-%m-%d %H:%M UTC')}"])
    ws.append([])
    
    # Top-level counts
    total_items = len(classification)
    in_screening = sum(1 for r in classification.values() if r["sources"])
    superseded_only = sum(1 for r in classification.values() if r["superseded"] and not r["sources"])
    superseded_and_screening = sum(1 for r in classification.values() if r["superseded"] and r["sources"])
    cross_db = sum(1 for r in classification.values() if len(r["sources"]) >= 2)
    single_source = sum(1 for r in classification.values() if len(r["sources"]) == 1)
    
    ws.append(["Counts", "Value"])
    style_header(ws, ws.max_row, 2)
    ws.append(["Total tracked items", total_items])
    ws.append(["Items in any screening collection", in_screening])
    ws.append(["Items in single source only", single_source])
    ws.append(["Items in 2+ sources (cross-database overlap)", cross_db])
    ws.append(["Items linked to Superseded (any membership)", sum(1 for r in classification.values() if r["superseded"])])
    ws.append(["Items linked to Superseded AND in screening", superseded_and_screening])
    ws.append(["Items in Superseded only (not in screening)", superseded_only])
    ws.append([])
    
    # Per-source counts
    ws.append(["Per-source item counts", "Items"])
    style_header(ws, ws.max_row, 2)
    src_counts = defaultdict(int)
    for r in classification.values():
        for s in r["sources"]:
            src_counts[s] += 1
    for s in sorted(src_counts):
        ws.append([f"source:{s}", src_counts[s]])
    ws.append([])
    
    # Cross-database overlap patterns
    ws.append(["Cross-database overlap patterns", "Items"])
    style_header(ws, ws.max_row, 2)
    pattern_counts = defaultdict(int)
    for r in classification.values():
        if len(r["sources"]) >= 2:
            pattern = " + ".join(sorted(r["sources"]))
            pattern_counts[pattern] += 1
    for pat in sorted(pattern_counts, key=lambda x: -pattern_counts[x]):
        ws.append([pat, pattern_counts[pat]])
    ws.append([])
    
    # Decision-disagreement patterns (when item is in 2+ sources with different decisions)
    ws.append(["Cross-source decision disagreements", "Items"])
    style_header(ws, ws.max_row, 2)
    disagree_count = 0
    for r in classification.values():
        if len(r["decisions"]) >= 2:
            decisions = set(r["decisions"].values())
            if len(decisions) >= 2:
                disagree_count += 1
    ws.append(["Items with conflicting decisions across sources", disagree_count])
    
    autosize(ws)

    # ------------------------------------------------------------
    # Sheet 2: Cross-DB overlaps
    # ------------------------------------------------------------
    ws = wb.create_sheet("Cross-DB")
    headers = ["item_key", "title", "sources", "decisions_per_source", "consistent", "superseded"]
    ws.append(headers)
    style_header(ws, 1, len(headers))
    cross_items = sorted(
        [(k, r) for k, r in classification.items() if len(r["sources"]) >= 2],
        key=lambda kv: (-len(kv[1]["sources"]), kv[1]["title"][:80])
    )
    for k, r in cross_items:
        decisions_str = "|".join(f"{s}:{d}" for s, d in sorted(r["decisions"].items()))
        decisions_set = set(r["decisions"].values())
        consistent = "yes" if len(decisions_set) <= 1 else "no"
        ws.append([
            k,
            r["title"][:200],
            "|".join(sorted(r["sources"])),
            decisions_str,
            consistent,
            "yes" if r["superseded"] else "",
        ])
    autosize(ws)

    # ------------------------------------------------------------
    # Sheet 3: Superseded
    # ------------------------------------------------------------
    ws = wb.create_sheet("Superseded")
    headers = ["item_key", "title", "superseded_collections", "also_in_sources", "decisions_per_source"]
    ws.append(headers)
    style_header(ws, 1, len(headers))
    super_items = sorted(
        [(k, r) for k, r in classification.items() if r["superseded"]],
        key=lambda kv: (-len(kv[1]["sources"]), kv[1]["title"][:80])
    )
    for k, r in super_items:
        ws.append([
            k,
            r["title"][:200],
            "|".join(r["superseded_collections"]),
            "|".join(sorted(r["sources"])) if r["sources"] else "(superseded only)",
            "|".join(f"{s}:{d}" for s, d in sorted(r["decisions"].items())),
        ])
    autosize(ws)

    # ------------------------------------------------------------
    # Sheet 4: All overlaps (anything in 2+ tracked collections, including superseded crossings)
    # ------------------------------------------------------------
    ws = wb.create_sheet("All overlaps")
    headers = ["item_key", "title", "n_sources", "sources", "decisions_per_source",
               "queue_only_in", "superseded", "superseded_collections"]
    ws.append(headers)
    style_header(ws, 1, len(headers))
    
    def overlap_count(r):
        # An "overlap" = total memberships across screening + superseded
        return len(r["sources"]) + (1 if r["superseded"] else 0)
    
    all_overlap = sorted(
        [(k, r) for k, r in classification.items() if overlap_count(r) >= 2],
        key=lambda kv: (-overlap_count(kv[1]), kv[1]["title"][:80])
    )
    for k, r in all_overlap:
        ws.append([
            k,
            r["title"][:200],
            len(r["sources"]),
            "|".join(sorted(r["sources"])),
            "|".join(f"{s}:{d}" for s, d in sorted(r["decisions"].items())),
            "|".join(sorted(r["queue_only_in"])),
            "yes" if r["superseded"] else "",
            "|".join(r["superseded_collections"]),
        ])
    autosize(ws)
    
    wb.save(path)
    log(f"  Wrote {path}")
    log(f"  Cross-DB overlaps: {len(cross_items)}")
    log(f"  Superseded items: {len(super_items)}")
    log(f"  All overlaps (any 2+ memberships): {len(all_overlap)}")

# ============================================================
# MAIN
# ============================================================
def main():
    parser = argparse.ArgumentParser(description=__doc__, formatter_class=argparse.RawDescriptionHelpFormatter)
    parser.add_argument("--apply", action="store_true",
                        help="Actually write tags via the read/write API key. Default is dry-run.")
    parser.add_argument("--report-only", action="store_true",
                        help="Skip tag analysis; only generate the overlap report.")
    args = parser.parse_args()

    if args.apply and args.report_only:
        print("ERROR: --apply and --report-only are mutually exclusive")
        sys.exit(1)

    log(f"=== Stage 1 Tag + Overlap Run ===")
    log(f"Mode: {'APPLY (writes tags)' if args.apply else 'DRY-RUN (no writes)' if not args.report_only else 'REPORT ONLY'}")

    collections = all_collections()
    classification = build_classification(collections)

    log(f"\nTotal tracked items: {len(classification)}")

    # Always generate the overlap report
    generate_overlap_report(classification, OUTPUT_REPORT)

    if args.report_only:
        log("\nReport-only mode; skipping tag plan.")
        return

    # Build tag plan
    plan = build_tag_plan(classification)
    write_tag_plan_csv(plan, TAG_PLAN_CSV)
    items_with_changes = sum(1 for p in plan if p["desired_new_tags"])
    total_new_tags = sum(len(p["desired_new_tags"]) for p in plan)
    log(f"\nTag plan written to {TAG_PLAN_CSV}")
    log(f"  Items needing new tags: {items_with_changes} of {len(plan)}")
    log(f"  Total new tags to add:  {total_new_tags}")

    # Sample
    log(f"\nSample of planned changes (first 10 items with changes):")
    sample = [p for p in plan if p["desired_new_tags"]][:10]
    for p in sample:
        log(f"  {p['key']}: + {p['desired_new_tags']}")

    if args.apply:
        log("\n--apply specified. Writing tags via read/write API key...")
        successes, failures = apply_tags(plan, ZOTERO_API_KEY_RW)
        log(f"\nFinal: {successes} items updated, {len(failures)} failures")
    else:
        log("\nDRY-RUN complete. To actually apply tags, re-run with --apply")
        log(f"Review {TAG_PLAN_CSV} and {OUTPUT_REPORT} before applying.")

if __name__ == "__main__":
    main()
